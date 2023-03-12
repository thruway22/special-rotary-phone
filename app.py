import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
import datetime
import calendar
import tempfile
import convertapi
from tempfile import NamedTemporaryFile

st.title('💰 TimesheetMaker')
st.write('An easy and quick timesheet maker for Baker boys.')

# simple, advanced = st.tabs(['Simple', 'Advanced'])
form = st.form('input_form')
left, middle, right = form.columns ([2, 1, 1])
employee_name = left.text_input('Name')
employee_id = middle.number_input('ID', step=1, min_value=0)
employee_rate = right.number_input('Day Rate (SAR)', min_value=0.0)
left, right = form.columns (2)
date_start = left.date_input('Date Start')
date_end = right.date_input('Date End')
left, right = form.columns (2)
rig_name = left.selectbox('Rig Name', ['', 'BCTD-4', 'BCTD-5'])
wstl_list = ['Ken Lynn', 'Pete Riley', 'Steve Baranyi', 'Ahmed Mansour']
wstl_list.sort()
wstl_name = right.selectbox('Wellsite Team Leader', [''] + wstl_list)
form.write('Caution: To produce a downloadable Excel file, an empty timesheet is loaded \
into your device private memory. All data are then processed on you local device private memory. \
However, to produce a PDF file, the excel file is sent to a server. ')
pdf_options = ["I'd like my data to stay privet, generate Excel file only",
                "I understand the caution,  generate Excel and PDF files"]
pdf = form.radio('pdf', pdf_options, label_visibility='collapsed')

submitted = form.form_submit_button('Generate File')

if submitted:
    if date_start > date_end:
        st.error('Starting date is later than ending date.')

    # elif date_start.month != date_end.month or date_start.year != date_end.year:
    #      st.error('At the moment, starting date and ending date must be in the same month/year. The possibility of making timesheets that spans over multiple months might be added later.')

    else:
        with st.spinner('Working on your timesheet...'):
            quick_accessor = employee_name
            if quick_accessor.startswith('#'):
                try:
                    employee_id = st.secrets[quick_accessor[1:]]['id']
                    employee_name = st.secrets[quick_accessor[1:]]['name']
                    employee_rate = st.secrets[quick_accessor[1:]]['rate']
                    rig_name = st.secrets[quick_accessor[1:]]['rig']
                except:
                     st.error('ERROR')
                     st.stop()

            # if date_start.month > date_end.month_end

            
            wb = load_workbook(filename=r'template.xlsx', read_only=False)
            ws = wb['timesheet']

            # remove footer
            ws.oddFooter.left.text = ''
            ws.oddFooter.center.text = ''
            ws.oddFooter.right.text = ''

            def next_month(input_date):
                new_year = input_date.year
                new_month = input_date.month + 1
                if new_month > 12:
                    new_year += 1
                    new_month -= 12
                    
                return new_month

            if date_end.month == next_month(date_start):
                wb.copy_worksheet(ws)
                ws2 = wb['timesheet Copy']
                ws2.title = '{}'.format(str(calendar.month_abbr[date_end.month].upper()) + str(date_start.year))
                
                month_start = 1
                month_end = calendar.monthrange(date_end.year, date_end.month)[1] + 1
                for day in range(month_start, month_end):
                    cell_a = 'A' + str(day + 1)
                    ws2[cell_a] = day

                shift_start = 1
                shift_end = date_end.day
                for shift in range(shift_start, shift_end):
                    cell_b = 'B' + str(shift + 1)
                    cell_d = 'D' + str(shift + 1)
                    ws[cell_b] = 'ARAMCO'
                    ws[cell_d] = rig_name.upper()

            
            month_start = 1
            month_end = calendar.monthrange(date_start.year, date_start.month)[1] + 1
            for day in range(month_start, month_end):
                cell_a = 'A' + str(day + 1)
                ws[cell_a] = day

            # def print_month_range(input_date, ws=ws):
            #     month_start = 1
            #     month_end = calendar.monthrange(input_date.year, input_date.month)[1] + 1
            #     for day in range(month_start, month_end):
            #         cell_a = 'A' + str(day + 1)
            #         globals()['ws'][cell_a] = day

            # print_month_range(date_start)

            shift_start = date_start.day
            shift_end = month_end if date_end.month > date_start.month else date_end.day + 1
            for shift in range(shift_start, shift_end):
                cell_b = 'B' + str(shift + 1)
                cell_d = 'D' + str(shift + 1)
                ws[cell_b] = 'ARAMCO'
                ws[cell_d] = rig_name.upper()
    
            ws['Q2']= employee_name.upper()
            ws['Q3']= employee_id
            ws['Q4']= 'KSA'
            ws['Q5']= str(calendar.month_abbr[date_start.month].upper()) + ' ' + str(date_start.year) # month year

            hitch = len(range(shift_start, shift_end)) # total shift days
            ws['O8']= hitch
            ws['Q8']= employee_rate
            ws['T8']= hitch * employee_rate
            ws['T19']= hitch * employee_rate

            ws['O22']= employee_name.upper()
            ws['O24']= wstl_name.upper()

            sheet_name = '{}'.format(str(calendar.month_abbr[date_start.month].upper()) + str(date_start.year))
            ws.title = sheet_name

            
            
            # if date_end.month - 1 == date_start.month and date_end.year == date_start.year:
            #     wb.copy_worksheet(ws)
            #     ws2 = wb['{} Copy'.format(sheet_name)]
            #     ws2.title = '{}'.format(str(calendar.month_abbr[date_end.month].upper()) + str(date_end.year))
            # wb.copy_worksheet(ws)
            # ws2 = wb['{} Copy'.format(sheet_name)]
            # ws2.title = "timesheet 2"
            output = BytesIO()
            wb.save(output)

            temp_file = tempfile.NamedTemporaryFile()
            with NamedTemporaryFile() as tmp:
                wb.save(tmp.name)
                output_tmp = BytesIO(tmp.read())
            

            #st.write('Hi {}, your hitch is {} days and total rate is {}SAR.'.format(employee_name, hitch, round(hitch * employee_rate)))
            st.success('Your timesheet has been successfully generated. Click on download button below.')
            output_file_name = 'TS-{}-{}'.format(
                employee_id, str(calendar.month_abbr[date_start.month].upper()) + str(date_start.year))

            file_name = 'Timesheet-{}'.format(employee_id)
            st.download_button(
                label='Download Excel File',
                data=output.getvalue(),
                file_name=output_file_name,
                #mime="application/vnd.ms-excel"
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            st.stop()
            #if pdf == pdf_options[1]:
            convertapi.api_secret = st.secrets['api_secret']

            content = output_tmp.getvalue()
            upload_io = convertapi.UploadIO(content, 'ts.xlsx')
            result = convertapi.convert('pdf', {
                'File': upload_io })
            saved_file = result.file.save(tempfile.gettempdir())

            with open(saved_file, "rb") as file:
                btn = st.download_button(
                        label="Download image",
                        data=file,
                        file_name="flower.pdf",
                        mime="application/octet-stream"
                    )


            st.stop()
           
            ###################
            ###################
            #st.stop()
            ###################
            ###################
            # output_html = BytesIO()
            # html = xlsx2html(output, b'output.html')
            ####################################
            out_stream = xlsx2html(output)
            out_stream.seek(0)

            st.download_button(
                label="Download Html Page",
                data=out_stream.read(),
                file_name="report.html",
                mime="application/octet-stream"
            )

            #st.write(out_stream.getvalue())

            st.stop()

            pdf_output = BytesIO()
            #result_file = open(pdf_output, "w+b")
            pisa_stat = pisa.CreatePDF(out_stream.getvalue(), dest=pdf_output)
            #pdf_output.close()
            #pdf_output.seek(0)

            st.download_button(
                label="Download pdf Page",
                data=pdf_output.getvalue(),
                file_name="report.pdf",
                mime="application/octet-stream")

            #st.code(out_stream.getvalue(), 'html')
            #st.stop()

            soup = BeautifulSoup(str(out_stream.getvalue()), 'html.parser')
            head = soup.find('head')
            soup.find('head')
            head.append(
                BeautifulSoup(
                    '<style>@page {size: A4 landscape; @frame content_frame {left: 2pt; width: 842pt; top: 2pt; height: 595pt;}} #col1 {width: 67.19999999999999;} #col2 {width: 157.44;} #col3 {width: 122.88;} #col4 {width: 120.0;} #col5 {width: 100.80000000000001;} #col6 {width: 67.19999999999999;} #col7 {width: 94.08;} #col8 {width: 67.19999999999999;} #col9 {width: 115.19999999999999;} #col10 {width: 100.80000000000001;} #col11 {width: 86.4;} #col12 {width: 67.19999999999999;} #col13 {width: 110.39999999999999;} #col14 {width: 67.19999999999999;} #col15 {width: 67.19999999999999;} #col16 {width: 38.400000000000006;} #col17 {width: 72.96000000000001;} #col18 {width: 70.08;} #col19 {width: 148.8;} #col20 {width: 100.80000000000001;} #col21 {width: 197.76;} #col22 {width: 84.48;} #col23 {width: 84.48;} #col24 {width: 84.48;} #col25 {width: 84.48;} #col26 {width: 84.48;} #col27 {width: 84.48;} #col28 {width: 84.48;}</style>', 'html.parser'))
            

            col = 1
            for i in soup.find_all('col'):
                del i['style']
                i['id'] = 'col'+ str(col)
                col += 1
                st.code(i, 'html')
            #st.stop()

            st.code(head, 'html')


            # BeautifulSoup('<tr>string</tr>', 'html.parser')

            # for i in soup.find_all('small'):
            #     if i.string :
            #         i.string.replace_with(i.string.replace(u'\xa0', '-'))
            # for i in soup.find_all('head'):
            #     if i.string :
            #         i.string.replace_with(
            #             '@page {size: a4 landscape;margin: 2cm;}'
            #             )
            # head.string.replace_with(
            #     '<meta charset="UTF-8"><title>Title</title> @page {size: letter landscape;margin: 2cm;}'
            #     )

            pdf_output2 = BytesIO()
            #result_file = open(pdf_output, "w+b")
            pisa_stat = pisa.CreatePDF(str(soup), dest=pdf_output2)
            #pdf_output.close()
            #pdf_output.seek(0)

            st.download_button(
                label="Download pdf Page 2",
                data=pdf_output2.getvalue(),
                file_name="lsreprot.pdf",
                mime="application/octet-stream")

            # pdf_out = BytesIO()
            # y = PdfWriter()
            # #y.addpage()
            # y.write(pdf_out)
            # # ws_range = ws.iter_rows()
            # # for row in ws_range:
            # #     s = ''
            # #     for cell in row:
            # #         if cell.value is None:
            # #             s = s
            # #         else:
            # #             s += str(cell.value) #.rjust(10) + ' '
            # #     pw.writeLine(s)
            # # pw.savePage()
            # # pw.close()

            # st.download_button(label="Export_Report",
            #             data=pdf_out.getvalue(),
            #             file_name="test.pdf",
            #             mime='application/octet-stream')

            st.write('test')

